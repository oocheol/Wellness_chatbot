
from tensorflow.keras.preprocessing.text import Tokenizer
from tensorflow.keras.preprocessing.sequence import pad_sequences
from tensorflow.keras.models import model_from_json
from transformers import BertTokenizer, TFBertModel
from openpyxl import load_workbook
import numpy as np
import tensorflow as tf

import pickle

tokenizer = Tokenizer()
with open('./model_data/tokenizer.pickle', 'rb') as handle:
    tokenizer = pickle.load(handle)


def convert_data(sentence):
    # 토큰화
    sentence = sentence.split(' ')
    # 정수화
    sentence = tokenizer.texts_to_sequences([sentence])
    # 패딩
    sentence = pad_sequences(sentence,
                             value=0,
                             padding='post',
                             maxlen=50)
    return sentence


def get_sentiment(sentence):
    with open("./model_data/rnn_sentiment_model.json", "r") as file:
        loaded_model_json = file.read()

    loaded_model = model_from_json(loaded_model_json)
    loaded_model.load_weights('./model_data/rnn_sentiment_model.h5')

    setence = convert_data(sentence)
    try:
        result = loaded_model.predict(setence)
        return result[0][0]
    except:
        return 0


# 파라미터 세팅
SEQ_LEN = 32      # 글자 길이

# @title BERT 사전학습 모델 불러오기 함수 정의
def create_sentiment_bert(CATEGORY_NUM):

    # 버트 pretrained 모델 로드
    model = TFBertModel.from_pretrained(
        "bert-base-multilingual-cased", from_pt=True)
    # 토큰 인풋, 마스크 인풋, 세그먼트 인풋 정의
    token_inputs = tf.keras.layers.Input(
        (SEQ_LEN,), dtype=tf.int32, name='input_word_ids')
    mask_inputs = tf.keras.layers.Input(
        (SEQ_LEN,), dtype=tf.int32, name='input_masks')
    segment_inputs = tf.keras.layers.Input(
        (SEQ_LEN,), dtype=tf.int32, name='input_segment')
    # 인풋이 [토큰, 마스크, 세그먼트]인 모델 정의
    bert_outputs = model([token_inputs, mask_inputs, segment_inputs])
    bert_outputs = bert_outputs[1]

    sentiment_first = tf.keras.layers.Dense(
        CATEGORY_NUM, activation='softmax', kernel_initializer=tf.keras.initializers.TruncatedNormal(0.02))(bert_outputs)
    sentiment_model = tf.keras.Model(
        [token_inputs, mask_inputs, segment_inputs], sentiment_first)

    sentiment_model.compile(optimizer=tf.keras.optimizers.Adam(lr=0.00001), loss=tf.keras.losses.SparseCategoricalCrossentropy(from_logits=False),
                            metrics=['sparse_categorical_accuracy'])
    return sentiment_model


# menu_intent
MENU_CATEGORY_NUM = 3
menu_intent_model = create_sentiment_bert(MENU_CATEGORY_NUM)
menu_intent_model.load_weights('./model_data/bert_menu_intent_model.h5')


SUB_CATEGORY_NUM = 2
sub_intent_model = create_sentiment_bert(SUB_CATEGORY_NUM)
sub_intent_model.load_weights('./model_data/bert_sub_intent_model.h5')

bert_tokenizer = BertTokenizer.from_pretrained('bert-base-multilingual-cased')

# @title [실행] BERT 모델 추론 관련 함수 정의
def mean_answer_label(*preds):
    preds_sum = np.zeros(preds[0].shape[0])
    for pred in preds:
        preds_sum += np.argmax(pred, axis=-1)
    return np.round(preds_sum/len(preds), 0).astype(int)


def sentence_convert_data(data):
    global bert_tokenizer

    tokens, masks, segments = [], [], []
    token = bert_tokenizer.encode(data, max_length=SEQ_LEN,
                             padding='max_length', truncation=True)

    num_zeros = token.count(0)
    mask = [1]*(SEQ_LEN-num_zeros) + [0]*num_zeros
    segment = [0]*SEQ_LEN

    tokens.append(token)
    segments.append(segment)
    masks.append(mask)

    tokens = np.array(tokens)
    masks = np.array(masks)
    segments = np.array(segments)
    return [tokens, masks, segments]


def predict(sentence, intent_model, cate_dict, LABEL_COLUMN):
    data_x = sentence_convert_data(sentence)
    message = intent_model.predict(data_x, batch_size=1)

    preds = str(mean_answer_label(message).item())
    return cate_dict[LABEL_COLUMN][preds]



def load_cate_dict(work_book):
    cate_dict = {
        "menu": {},
        "sub": {}
    }
    for menu_index, work_sheet in enumerate(work_book.worksheets):
        sheet_name = work_sheet.title
        cate_dict['menu'][str(menu_index)] = sheet_name

        sheet = work_book[sheet_name]
        for column_index, column in enumerate(sheet.columns):
            for index, row in enumerate(column):
                if index is 0:
                    cate_dict['sub'][str(column_index)] = row.value
    return cate_dict


work_book = load_workbook('./model_data/coffee.xlsx', data_only=True)
cate_dict = load_cate_dict(work_book)


def get_menu_intent(message):
    intent = predict(message, menu_intent_model, cate_dict, "menu")
    return intent

def get_sub_intent(message):
    intent = predict(message, sub_intent_model, cate_dict, "sub")
    return intent



work_book = load_workbook('./model_data/coffee_response.xlsx', data_only=True)
work_sheet = work_book.active

def get_response(menu_intent, sub_intent):
    try:
        for index, row in enumerate(work_sheet.rows):
            menu = row[0].value
            sub = row[1].value
            response = row[2].value

            if menu_intent in menu and sub_intent in sub:
                return response
        return "주문을 다시 해 주세요."
    except:
        return "주문을 다시 해 주세요."
